"""
Vistas REST para el módulo Beneficios de Salud (SIGA).
"""
import hashlib
import os
from decimal import Decimal

from django.conf import settings
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    ArchivoRecibido, BeneficioSalud, ErrorProcesamiento,
    PoliticaPrepagada, PensionadoPrepagada, AuxilioExterno,
    PlanillaCalculo, DetalleCalculo,
)
from .serializers import (
    ArchivoRecibidoListSerializer,
    ArchivoRecibidoDetailSerializer,
    BeneficioSaludSerializer,
    PoliticaPrepagadaSerializer,
    PensionadoPrepagadaSerializer,
    AuxilioExternoSerializer,
    PlanillaCalculoListSerializer,
    PlanillaCalculoDetailSerializer,
    DetalleCalculoSerializer,
)
from .services.detector import detectar_proveedor
from .services.reader_excel import leer_excel
from .services.axa_adapter import adaptar_axa
from .services.colsanitas_adapter import adaptar_colsanitas
from .services.validator import validar_registros
from .services import prepagada_service


def _sha256_archivo(file_obj) -> str:
    """Calcula el hash SHA256 de un objeto archivo."""
    sha256 = hashlib.sha256()
    file_obj.seek(0)
    for chunk in iter(lambda: file_obj.read(8192), b''):
        sha256.update(chunk)
    file_obj.seek(0)
    return sha256.hexdigest()


def _guardar_archivo(file_obj, proveedor: str, nombre_archivo: str) -> str:
    """
    Guarda el archivo en storage/landing/{proveedor}/{nombre_archivo}.
    Retorna la ruta absoluta donde fue guardado.
    """
    directorio_proveedor = {
        'axa': 'axa_colpatria',
        'colsanitas': 'colsanitas',
        'desconocido': 'desconocido',
    }.get(proveedor, 'desconocido')

    destino_dir = os.path.join(settings.MEDIA_ROOT, directorio_proveedor)
    os.makedirs(destino_dir, exist_ok=True)

    ruta_destino = os.path.join(destino_dir, nombre_archivo)

    file_obj.seek(0)
    with open(ruta_destino, 'wb') as f:
        for chunk in file_obj.chunks():
            f.write(chunk)

    return ruta_destino


class UploadView(APIView):
    """
    POST /api/beneficios-salud/upload/
    Recibe un archivo Excel via multipart form y lo procesa.
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, *args, **kwargs):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'error': 'Se requiere el campo "archivo" con el archivo Excel.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        usuario = 'anonimo'
        if request.user and request.user.is_authenticated:
            usuario = request.user.username
        elif request.data.get('usuario'):
            usuario = str(request.data.get('usuario'))

        nombre_archivo = archivo.name
        hash_archivo = _sha256_archivo(archivo)

        # Detección inicial del proveedor por nombre de archivo
        proveedor = detectar_proveedor(nombre_archivo)

        # Guardar archivo en disco
        try:
            ruta_archivo = _guardar_archivo(archivo, proveedor, nombre_archivo)
        except Exception as exc:
            return Response(
                {'error': f'No se pudo guardar el archivo: {exc}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Crear registro ArchivoRecibido
        archivo_obj = ArchivoRecibido.objects.create(
            proveedor=proveedor,
            nombre_archivo=nombre_archivo,
            ruta_archivo=ruta_archivo,
            estado_procesamiento='RECIBIDO',
            hash_archivo=hash_archivo,
            usuario_carga=usuario,
        )

        try:
            # Marcar como PROCESANDO
            archivo_obj.estado_procesamiento = 'PROCESANDO'
            archivo_obj.save(update_fields=['estado_procesamiento'])

            # Leer Excel
            df, metadatos = leer_excel(ruta_archivo, proveedor)

            # Refinar detección por columnas si el proveedor era desconocido
            if proveedor == 'desconocido':
                proveedor = detectar_proveedor(nombre_archivo, list(df.columns))
                archivo_obj.proveedor = proveedor

            # Actualizar metadatos en el registro
            archivo_obj.numero_contrato = metadatos.get('numero_contrato', '')
            archivo_obj.periodo_facturacion = metadatos.get('periodo_facturacion', '')

            # Adaptar al esquema unificado
            if proveedor == 'axa':
                df_unificado = adaptar_axa(df, metadatos)
            elif proveedor == 'colsanitas':
                df_unificado = adaptar_colsanitas(df, metadatos)
            else:
                raise ValueError(
                    f'Proveedor no reconocido: "{proveedor}". '
                    'El nombre del archivo debe contener "AXA" o "COLSANITAS".'
                )

            # Añadir archivo_origen al dataframe
            df_unificado['archivo_origen'] = nombre_archivo

            total_registros = len(df_unificado)
            archivo_obj.total_registros = total_registros
            archivo_obj.save(update_fields=['proveedor', 'numero_contrato', 'periodo_facturacion', 'total_registros'])

            # Validar registros
            registros_ok, errores_val = validar_registros(df_unificado, archivo_obj.id)

            # Inserción bulk de beneficios
            beneficios_bulk = [BeneficioSalud(**r) for r in registros_ok]
            BeneficioSalud.objects.bulk_create(beneficios_bulk, batch_size=500)

            # Inserción bulk de errores
            errores_bulk = [ErrorProcesamiento(**e) for e in errores_val]
            ErrorProcesamiento.objects.bulk_create(errores_bulk, batch_size=500)

            registros_procesados = len(registros_ok)
            # Errores fatales = registros que no se insertaron (no incluye advertencias)
            registros_con_error = total_registros - registros_procesados

            # Actualizar estadísticas
            archivo_obj.registros_procesados = registros_procesados
            archivo_obj.registros_con_error = registros_con_error
            archivo_obj.estado_procesamiento = 'PROCESADO'
            archivo_obj.save(update_fields=[
                'registros_procesados',
                'registros_con_error',
                'estado_procesamiento',
            ])

            return Response({
                'archivo_id': archivo_obj.id,
                'proveedor': archivo_obj.proveedor,
                'total_registros': total_registros,
                'registros_procesados': registros_procesados,
                'registros_con_error': registros_con_error,
                'estado': 'PROCESADO',
            }, status=status.HTTP_201_CREATED)

        except Exception as exc:
            archivo_obj.estado_procesamiento = 'ERROR'
            archivo_obj.save(update_fields=['estado_procesamiento'])
            return Response(
                {
                    'error': str(exc),
                    'archivo_id': archivo_obj.id,
                    'estado': 'ERROR',
                },
                status=status.HTTP_422_UNPROCESSABLE_ENTITY
            )


class ArchivoListView(APIView):
    """
    GET /api/beneficios-salud/archivos/
    Lista todos los ArchivoRecibido.
    """

    def get(self, request, *args, **kwargs):
        qs = ArchivoRecibido.objects.all()

        proveedor = request.query_params.get('proveedor')
        if proveedor:
            qs = qs.filter(proveedor=proveedor)

        estado = request.query_params.get('estado')
        if estado:
            qs = qs.filter(estado_procesamiento=estado)

        serializer = ArchivoRecibidoListSerializer(qs, many=True)
        return Response(serializer.data)


class ArchivoDetailView(APIView):
    """
    GET /api/beneficios-salud/archivos/{id}/
    Detalle de un archivo con sus errores.
    """

    def get(self, request, pk, *args, **kwargs):
        try:
            archivo = ArchivoRecibido.objects.get(pk=pk)
        except ArchivoRecibido.DoesNotExist:
            return Response(
                {'error': f'Archivo con id={pk} no encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = ArchivoRecibidoDetailSerializer(archivo)
        return Response(serializer.data)


class BeneficioListView(APIView):
    """
    GET /api/beneficios-salud/beneficios/
    Lista BeneficioSalud con filtros opcionales: ?archivo_id=X&proveedor=X&cedula=X
    """

    def get(self, request, *args, **kwargs):
        qs = BeneficioSalud.objects.select_related('archivo').all()

        archivo_id = request.query_params.get('archivo_id')
        if archivo_id:
            qs = qs.filter(archivo_id=archivo_id)

        proveedor = request.query_params.get('proveedor')
        if proveedor:
            qs = qs.filter(proveedor=proveedor)

        cedula = request.query_params.get('cedula')
        if cedula:
            qs = qs.filter(cedula=cedula)

        estado_validacion = request.query_params.get('estado_validacion')
        if estado_validacion:
            qs = qs.filter(estado_validacion=estado_validacion)

        serializer = BeneficioSaludSerializer(qs, many=True)
        return Response(serializer.data)


class ExportarExcelView(APIView):
    """
    GET /api/beneficios-salud/exportar/?archivo_id=X
    Exporta beneficios de salud a un archivo Excel (.xlsx) con tres hojas:
    Consolidado, AXA Colpatria, Colsanitas.
    Si se pasa archivo_id, exporta ese archivo.
    Si no, exporta el último archivo PROCESADO de cada proveedor.
    """

    COLUMNAS = [
        'cedula', 'nombre', 'parentesco', 'sub_contrato', 'cedula_titular',
        'proveedor', 'tipo_plan', 'valor_base', 'descuento', 'iva',
        'valor_total', 'fecha_corte', 'numero_contrato', 'periodo_facturacion',
        'estado_validacion',
    ]

    def get(self, request, *args, **kwargs):
        from io import BytesIO
        from datetime import date
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from django.http import HttpResponse
        from django.db.models import Max

        archivo_id = request.query_params.get('archivo_id')

        qs = BeneficioSalud.objects.select_related('archivo').filter(
            archivo__estado_procesamiento='PROCESADO'
        )

        if archivo_id:
            qs = qs.filter(archivo_id=archivo_id)
        else:
            # Obtener el max(id) de ArchivoRecibido por proveedor entre los procesados
            max_ids = (
                ArchivoRecibido.objects
                .filter(estado_procesamiento='PROCESADO')
                .values('proveedor')
                .annotate(max_id=Max('id'))
                .values_list('max_id', flat=True)
            )
            qs = qs.filter(archivo_id__in=list(max_ids))

        qs = qs.order_by('proveedor', 'nombre')

        registros_raw = list(qs.values(
            'cedula', 'nombre', 'parentesco', 'sub_contrato', 'cedula_titular',
            'proveedor', 'tipo_plan', 'valor_base', 'descuento', 'iva',
            'valor_total', 'fecha_corte', 'numero_contrato',
            'estado_validacion', 'archivo__nombre_archivo',
            'archivo__periodo_facturacion',
        ))
        # Normalizar para que el template use 'periodo_facturacion'
        registros = []
        for r in registros_raw:
            r['periodo_facturacion'] = r.pop('archivo__periodo_facturacion', '') or ''
            registros.append(r)

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color='00853f', end_color='00853f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        warn_fill   = PatternFill(start_color='FFFDE7', end_color='FFFDE7', fill_type='solid')

        def _escribir_hoja(ws, filas, incluir_origen=False):
            cols = self.COLUMNAS + (['archivo_origen'] if incluir_origen else [])
            # Encabezado
            for col_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
            # Datos
            for row_idx, reg in enumerate(filas, start=2):
                es_advertencia = reg.get('estado_validacion') == 'ADVERTENCIA'
                for col_idx, col_name in enumerate(cols, start=1):
                    if col_name == 'archivo_origen':
                        valor = reg.get('archivo__nombre_archivo', '')
                    else:
                        valor = reg.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                    if es_advertencia:
                        cell.fill = warn_fill

        # Hoja Consolidado
        ws_consolidado = wb.active
        ws_consolidado.title = 'Consolidado'
        _escribir_hoja(ws_consolidado, registros, incluir_origen=True)

        # Hoja AXA Colpatria
        ws_axa = wb.create_sheet(title='AXA Colpatria')
        filas_axa = [r for r in registros if r.get('proveedor') == 'axa']
        _escribir_hoja(ws_axa, filas_axa, incluir_origen=False)

        # Hoja Colsanitas
        ws_col = wb.create_sheet(title='Colsanitas')
        filas_col = [r for r in registros if r.get('proveedor') == 'colsanitas']
        _escribir_hoja(ws_col, filas_col, incluir_origen=False)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fecha_hoy = date.today().strftime('%Y%m%d')
        nombre_archivo = f'SIGA_Beneficios_{fecha_hoy}.xlsx'

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response


class NovedadesView(APIView):
    """
    GET /api/beneficios-salud/novedades/?archivo_nuevo=X&archivo_anterior=Y
    Compara dos archivos de BeneficioSalud y devuelve las novedades:
    nuevos afiliados, retirados y cambios de valor de cuota.
    """

    def get(self, request, *args, **kwargs):
        archivo_nuevo_id = request.query_params.get('archivo_nuevo')
        archivo_anterior_id = request.query_params.get('archivo_anterior')

        if not archivo_nuevo_id or not archivo_anterior_id:
            return Response(
                {'error': 'Se requieren los parámetros "archivo_nuevo" y "archivo_anterior".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            archivo_nuevo = ArchivoRecibido.objects.get(pk=archivo_nuevo_id)
        except ArchivoRecibido.DoesNotExist:
            return Response(
                {'error': f'Archivo con id={archivo_nuevo_id} no encontrado.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            archivo_anterior = ArchivoRecibido.objects.get(pk=archivo_anterior_id)
        except ArchivoRecibido.DoesNotExist:
            return Response(
                {'error': f'Archivo con id={archivo_anterior_id} no encontrado.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        warning = None
        if archivo_nuevo.proveedor != archivo_anterior.proveedor:
            warning = (
                f'Los archivos son de distinto proveedor: '
                f'"{archivo_nuevo.proveedor}" vs "{archivo_anterior.proveedor}".'
            )

        # Cargar registros de cada archivo como dicts por cédula
        qs_nuevo = BeneficioSalud.objects.filter(archivo_id=archivo_nuevo_id).values(
            'cedula', 'nombre', 'parentesco', 'valor_total'
        )
        qs_anterior = BeneficioSalud.objects.filter(archivo_id=archivo_anterior_id).values(
            'cedula', 'nombre', 'parentesco', 'valor_total'
        )

        mapa_nuevo    = {r['cedula']: r for r in qs_nuevo}
        mapa_anterior = {r['cedula']: r for r in qs_anterior}

        cedulas_nuevo    = set(mapa_nuevo.keys())
        cedulas_anterior = set(mapa_anterior.keys())

        cedulas_nuevos    = cedulas_nuevo - cedulas_anterior
        cedulas_retirados = cedulas_anterior - cedulas_nuevo
        cedulas_comunes   = cedulas_nuevo & cedulas_anterior

        nuevos = [
            {
                'cedula':     c,
                'nombre':     mapa_nuevo[c]['nombre'],
                'parentesco': mapa_nuevo[c]['parentesco'],
                'valor_total': mapa_nuevo[c]['valor_total'],
            }
            for c in sorted(cedulas_nuevos)
        ]

        retirados = [
            {
                'cedula':     c,
                'nombre':     mapa_anterior[c]['nombre'],
                'parentesco': mapa_anterior[c]['parentesco'],
                'valor_total': mapa_anterior[c]['valor_total'],
            }
            for c in sorted(cedulas_retirados)
        ]

        cambios_valor = []
        sin_cambios_count = 0
        for c in cedulas_comunes:
            vn = mapa_nuevo[c]['valor_total'] or 0
            va = mapa_anterior[c]['valor_total'] or 0
            try:
                vn_f = float(vn)
                va_f = float(va)
            except (TypeError, ValueError):
                vn_f = 0.0
                va_f = 0.0
            if abs(vn_f - va_f) > 1.0:
                cambios_valor.append({
                    'cedula':        c,
                    'nombre':        mapa_nuevo[c]['nombre'],
                    'parentesco':    mapa_nuevo[c]['parentesco'],
                    'valor_anterior': va,
                    'valor_nuevo':    vn,
                    'diferencia':     round(vn_f - va_f, 2),
                })
            else:
                sin_cambios_count += 1

        data = {
            'archivo_nuevo': {
                'id':                  archivo_nuevo.id,
                'nombre_archivo':      archivo_nuevo.nombre_archivo,
                'proveedor':           archivo_nuevo.proveedor,
                'periodo_facturacion': archivo_nuevo.periodo_facturacion,
            },
            'archivo_anterior': {
                'id':                  archivo_anterior.id,
                'nombre_archivo':      archivo_anterior.nombre_archivo,
                'proveedor':           archivo_anterior.proveedor,
                'periodo_facturacion': archivo_anterior.periodo_facturacion,
            },
            'resumen': {
                'nuevos':        len(nuevos),
                'retirados':     len(retirados),
                'cambios_valor': len(cambios_valor),
                'sin_cambios':   sin_cambios_count,
            },
            'nuevos':        nuevos,
            'retirados':     retirados,
            'cambios_valor': cambios_valor,
        }

        if warning:
            data['warning'] = warning

        return Response(data)


class DashboardView(APIView):
    """
    GET /api/beneficios-salud/dashboard/
    Resumen ejecutivo para el dashboard visual de SIGA.
    """

    PARENTESCO_LABEL = {
        'T': 'Titular', 'CO': 'Cónyuge', 'HI': 'Hijo(a)',
        'P': 'Padres', 'OT': 'Otro', '': 'Sin especificar',
    }

    def get(self, request, *args, **kwargs):
        from django.db.models import Max, Sum, Count

        # ── Último archivo procesado por proveedor ──────────────────────────
        ultimos_ids = (
            ArchivoRecibido.objects
            .filter(estado_procesamiento='PROCESADO')
            .values('proveedor')
            .annotate(max_id=Max('id'))
            .values_list('max_id', flat=True)
        )
        ultimos_archivos = ArchivoRecibido.objects.filter(id__in=list(ultimos_ids))

        ultimos_periodos = []
        for arch in ultimos_archivos.order_by('proveedor'):
            agg = BeneficioSalud.objects.filter(archivo=arch).aggregate(
                total_beneficiarios=Count('id'),
                valor_total=Sum('valor_total'),
            )
            ultimos_periodos.append({
                'proveedor':           arch.proveedor,
                'periodo':             arch.periodo_facturacion or '—',
                'numero_contrato':     arch.numero_contrato or '—',
                'archivo_id':          arch.id,
                'nombre_archivo':      arch.nombre_archivo,
                'fecha_procesamiento': arch.fecha_recepcion.date().isoformat() if arch.fecha_recepcion else None,
                'total_beneficiarios': agg['total_beneficiarios'] or 0,
                'valor_total':         float(agg['valor_total'] or 0),
                'registros_procesados': arch.registros_procesados,
                'registros_con_error':  arch.registros_con_error,
            })

        # ── Distribución por parentesco (sobre últimos archivos) ────────────
        qs_ultimos = BeneficioSalud.objects.filter(archivo_id__in=list(ultimos_ids))

        dist_parentesco_raw = (
            qs_ultimos
            .values('parentesco')
            .annotate(count=Count('id'), valor_total=Sum('valor_total'))
            .order_by('-count')
        )
        distribucion_parentesco = []
        for row in dist_parentesco_raw:
            p = (row['parentesco'] or '').strip().upper()
            distribucion_parentesco.append({
                'parentesco':  row['parentesco'] or '',
                'label':       self.PARENTESCO_LABEL.get(p, row['parentesco'] or 'Sin especificar'),
                'count':       row['count'],
                'valor_total': float(row['valor_total'] or 0),
            })

        # ── Distribución de valor por proveedor (últimos archivos) ──────────
        dist_proveedor = []
        for row in (
            qs_ultimos
            .values('proveedor')
            .annotate(count=Count('id'), valor_total=Sum('valor_total'))
            .order_by('proveedor')
        ):
            dist_proveedor.append({
                'proveedor':   row['proveedor'],
                'count':       row['count'],
                'valor_total': float(row['valor_total'] or 0),
            })

        # ── Evolución histórica (valor_total por archivo procesado) ─────────
        evolucion = []
        for arch in (
            ArchivoRecibido.objects
            .filter(estado_procesamiento='PROCESADO')
            .order_by('id')
        ):
            agg = BeneficioSalud.objects.filter(archivo=arch).aggregate(
                valor_total=Sum('valor_total'),
                count=Count('id'),
            )
            evolucion.append({
                'archivo_id':   arch.id,
                'proveedor':    arch.proveedor,
                'periodo':      arch.periodo_facturacion or arch.nombre_archivo[:30],
                'valor_total':  float(agg['valor_total'] or 0),
                'beneficiarios': agg['count'] or 0,
                'fecha':        arch.fecha_recepcion.date().isoformat() if arch.fecha_recepcion else None,
            })

        # ── Consolidado global ───────────────────────────────────────────────
        total_archivos_procesados = ArchivoRecibido.objects.filter(
            estado_procesamiento='PROCESADO'
        ).count()

        consolidado_agg = qs_ultimos.aggregate(
            valor_total=Sum('valor_total'),
            beneficiarios=Count('id'),
        )

        return Response({
            'ultimos_periodos':       ultimos_periodos,
            'distribucion_parentesco': distribucion_parentesco,
            'distribucion_proveedor':  dist_proveedor,
            'evolucion':               evolucion,
            'consolidado': {
                'total_archivos_procesados': total_archivos_procesados,
                'beneficiarios_ultimo_periodo': consolidado_agg['beneficiarios'] or 0,
                'valor_total_ultimo_periodo':   float(consolidado_agg['valor_total'] or 0),
                'proveedores_activos':          len(ultimos_periodos),
            },
        })


# ===========================================================================
# Medicina Prepagada views
# ===========================================================================

class CruceView(APIView):
    """
    GET /api/beneficios-salud/cruce/?periodo=MMYYYY
    Retorna registros de v_cruce para el periodo indicado,
    junto con la lista de periodos disponibles.
    """

    def get(self, request, *args, **kwargs):
        try:
            periodos = prepagada_service.get_periodos_disponibles()
        except RuntimeError as e:
            return Response({'error': str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        periodo = request.query_params.get('periodo')
        if not periodo:
            return Response({
                'periodos_disponibles': periodos,
                'cruce': [],
            })

        try:
            cruce = prepagada_service.get_cruce_periodo(periodo)
        except RuntimeError as e:
            return Response({'error': str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        return Response({
            'periodo': periodo,
            'periodos_disponibles': periodos,
            'total': len(cruce),
            'cruce': cruce,
        })


class PoliticaView(APIView):
    """
    GET  /api/beneficios-salud/politica/ → lista todas las políticas (más reciente primero)
    POST /api/beneficios-salud/politica/ → crea una nueva política
    """

    def get(self, request, *args, **kwargs):
        qs = PoliticaPrepagada.objects.all()
        serializer = PoliticaPrepagadaSerializer(qs, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        data = request.data.copy()
        if not data.get('creada_por') and request.user and request.user.is_authenticated:
            data['creada_por'] = request.user.username
        serializer = PoliticaPrepagadaSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PoliticaDetailView(APIView):
    """
    GET /api/beneficios-salud/politica/<pk>/
    PUT /api/beneficios-salud/politica/<pk>/
    """

    def _get_object(self, pk):
        try:
            return PoliticaPrepagada.objects.get(pk=pk)
        except PoliticaPrepagada.DoesNotExist:
            return None

    def get(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Política con id={pk} no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(PoliticaPrepagadaSerializer(obj).data)

    def put(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Política con id={pk} no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PoliticaPrepagadaSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PensionadosView(APIView):
    """
    GET  /api/beneficios-salud/pensionados/ → lista todos (activos por defecto)
    POST /api/beneficios-salud/pensionados/ → crea nuevo
    """

    def get(self, request, *args, **kwargs):
        qs = PensionadoPrepagada.objects.all()
        solo_activos = request.query_params.get('activo', '').lower()
        if solo_activos in ('1', 'true', 'yes'):
            qs = qs.filter(activo=True)
        serializer = PensionadoPrepagadaSerializer(qs, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        serializer = PensionadoPrepagadaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PensionadoDetailView(APIView):
    """
    GET    /api/beneficios-salud/pensionados/<pk>/
    PUT    /api/beneficios-salud/pensionados/<pk>/
    DELETE /api/beneficios-salud/pensionados/<pk>/
    """

    def _get_object(self, pk):
        try:
            return PensionadoPrepagada.objects.get(pk=pk)
        except PensionadoPrepagada.DoesNotExist:
            return None

    def get(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Pensionado con id={pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(PensionadoPrepagadaSerializer(obj).data)

    def put(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Pensionado con id={pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PensionadoPrepagadaSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Pensionado con id={pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AuxilioExternoView(APIView):
    """
    GET  /api/beneficios-salud/auxilio-externo/ → lista todos
    POST /api/beneficios-salud/auxilio-externo/ → crea nuevo
    """

    def get(self, request, *args, **kwargs):
        qs = AuxilioExterno.objects.all()
        solo_activos = request.query_params.get('activo', '').lower()
        if solo_activos in ('1', 'true', 'yes'):
            qs = qs.filter(activo=True)
        serializer = AuxilioExternoSerializer(qs, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        serializer = AuxilioExternoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class AuxilioExternoDetailView(APIView):
    """
    GET    /api/beneficios-salud/auxilio-externo/<pk>/
    PUT    /api/beneficios-salud/auxilio-externo/<pk>/
    DELETE /api/beneficios-salud/auxilio-externo/<pk>/
    """

    def _get_object(self, pk):
        try:
            return AuxilioExterno.objects.get(pk=pk)
        except AuxilioExterno.DoesNotExist:
            return None

    def get(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Auxilio externo con id={pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(AuxilioExternoSerializer(obj).data)

    def put(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Auxilio externo con id={pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = AuxilioExternoSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, *args, **kwargs):
        obj = self._get_object(pk)
        if obj is None:
            return Response({'error': f'Auxilio externo con id={pk} no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class PlanillaCalcularView(APIView):
    """
    POST /api/beneficios-salud/planilla/calcular/
    Body: {"periodo": "MMYYYY", "politica_id": X (opcional)}
    Calcula la planilla 80/20, guarda PlanillaCalculo + DetalleCalculo y retorna el resultado.
    """

    def post(self, request, *args, **kwargs):
        periodo = request.data.get('periodo')
        if not periodo:
            return Response({'error': 'El campo "periodo" es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        politica_id = request.data.get('politica_id')
        if politica_id:
            try:
                politica = PoliticaPrepagada.objects.get(pk=politica_id)
            except PoliticaPrepagada.DoesNotExist:
                return Response(
                    {'error': f'Política con id={politica_id} no encontrada.'},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            politica = PoliticaPrepagada.objects.order_by('-vigente_desde').first()
            if politica is None:
                return Response(
                    {'error': 'No existe ninguna política definida. Cree una política antes de calcular.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        try:
            resultados = prepagada_service.calcular_planilla(periodo, politica)
        except RuntimeError as e:
            return Response({'error': str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        # Calcular totales sobre registros liquidables. Los bloqueados quedan
        # visibles en el detalle, pero no suman aportes de empresa/empleado.
        liquidables = [
            r for r in resultados
            if r['estado_elegibilidad'] in ('ELEGIBLE_80_20', 'PENSIONADO_100')
        ]
        empleados_elegibles = [
            r for r in resultados
            if r['estado_elegibilidad'] == 'ELEGIBLE_80_20'
        ]
        total_empresa = sum(r['valor_empresa'] for r in liquidables)
        total_empleado = sum(r['valor_empleado'] for r in liquidables)
        total_gravable = sum(r['apoyo_gravable'] for r in liquidables)
        total_no_gravable = sum(r['apoyo_no_gravable'] for r in liquidables)

        generada_por = 'anonimo'
        if request.user and request.user.is_authenticated:
            generada_por = request.user.username

        planilla = PlanillaCalculo.objects.create(
            periodo=periodo,
            politica=politica,
            total_empleados=len(empleados_elegibles),
            total_empresa=round(total_empresa, 2),
            total_empleado=round(total_empleado, 2),
            total_gravable=round(total_gravable, 2),
            total_no_gravable=round(total_no_gravable, 2),
            generada_por=generada_por,
        )

        detalles_bulk = [
            DetalleCalculo(planilla=planilla, **r)
            for r in resultados
        ]
        DetalleCalculo.objects.bulk_create(detalles_bulk, batch_size=500)

        serializer = PlanillaCalculoDetailSerializer(planilla)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class PlanillaListView(APIView):
    """
    GET /api/beneficios-salud/planilla/
    Lista todas las planillas de cálculo.
    """

    def get(self, request, *args, **kwargs):
        qs = PlanillaCalculo.objects.select_related('politica').all()
        periodo = request.query_params.get('periodo')
        if periodo:
            qs = qs.filter(periodo=periodo)
        serializer = PlanillaCalculoListSerializer(qs, many=True)
        return Response(serializer.data)


class PlanillaDetailView(APIView):
    """
    GET /api/beneficios-salud/planilla/<pk>/
    Detalle de planilla con todos sus registros de DetalleCalculo.
    """

    def get(self, request, pk, *args, **kwargs):
        try:
            planilla = PlanillaCalculo.objects.select_related('politica').get(pk=pk)
        except PlanillaCalculo.DoesNotExist:
            return Response({'error': f'Planilla con id={pk} no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PlanillaCalculoDetailSerializer(planilla)
        return Response(serializer.data)


class PlanillaExportarView(APIView):
    """
    GET /api/beneficios-salud/planilla/<pk>/exportar/
    Exporta la planilla a Excel (.xlsx).
    Hoja 1: "Planilla 80-20" — todos los empleados.
    Hoja 2: "Apoyo Gravable" — solo los que tienen apoyo_gravable > 0.
    """

    COLUMNAS = [
        ('cedula', 'Cédula'),
        ('nombre_en_kactus', 'Nombre Kactus'),
        ('eps', 'EPS'),
        ('num_beneficiarios', 'N° Beneficiarios'),
        ('total_familia', 'Total Familia'),
        ('tipo_persona', 'Tipo Persona'),
        ('estado_elegibilidad', 'Elegibilidad'),
        ('motivo_elegibilidad', 'Motivo Elegibilidad'),
        ('porcentaje_empresa_aplicado', '% Empresa'),
        ('porcentaje_empleado_aplicado', '% Empleado'),
        ('valor_empresa', 'Valor Empresa'),
        ('valor_empleado', 'Valor Empleado'),
        ('valor_no_cubierto', 'Valor No Cubierto'),
        ('apoyo_no_gravable', 'Apoyo No Gravable'),
        ('apoyo_gravable', 'Apoyo Gravable'),
        ('estado_cruce', 'Estado Cruce'),
    ]

    def get(self, request, pk, *args, **kwargs):
        from io import BytesIO
        from datetime import date
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from django.http import HttpResponse

        try:
            planilla = PlanillaCalculo.objects.select_related('politica').get(pk=pk)
        except PlanillaCalculo.DoesNotExist:
            return Response({'error': f'Planilla con id={pk} no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

        detalles = list(planilla.detalles.all().values(
            'cedula', 'nombre_en_kactus', 'eps', 'num_beneficiarios',
            'total_familia', 'valor_empresa', 'valor_empleado',
            'valor_no_cubierto', 'apoyo_no_gravable', 'apoyo_gravable',
            'estado_cruce', 'tipo_persona', 'estado_elegibilidad',
            'motivo_elegibilidad', 'porcentaje_empresa_aplicado',
            'porcentaje_empleado_aplicado',
        ))

        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color='00853f', end_color='00853f', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        def _escribir_hoja(ws, filas):
            # Encabezado
            for col_idx, (campo, titulo) in enumerate(self.COLUMNAS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=titulo)
                cell.fill = header_fill
                cell.font = header_font
            # Datos
            for row_idx, reg in enumerate(filas, start=2):
                tiene_gravable = (reg.get('apoyo_gravable') or 0) > 0
                for col_idx, (campo, _) in enumerate(self.COLUMNAS, start=1):
                    valor = reg.get(campo)
                    # Convertir Decimal a float para openpyxl
                    if hasattr(valor, 'is_finite'):
                        valor = float(valor)
                    cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                    if tiene_gravable:
                        cell.fill = yellow_fill

        # Hoja 1 — Planilla completa
        ws1 = wb.active
        ws1.title = 'Planilla 80-20'
        _escribir_hoja(ws1, detalles)

        # Hoja 2 — Solo apoyo gravable
        ws2 = wb.create_sheet(title='Apoyo Gravable')
        gravables = [r for r in detalles if (r.get('apoyo_gravable') or 0) > 0]
        _escribir_hoja(ws2, gravables)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fecha_hoy = date.today().strftime('%Y%m%d')
        nombre_archivo = f'SIGA_Prepagada_{planilla.periodo}_{fecha_hoy}.xlsx'

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response


class CausacionView(APIView):
    """
    GET /api/beneficios-salud/causacion/?periodo=MMYYYY
    Resumen por EPS de la planilla del periodo indicado.
    """

    def get(self, request, *args, **kwargs):
        periodo = request.query_params.get('periodo')
        if not periodo:
            return Response({'error': 'El parámetro "periodo" es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        # Obtener la planilla más reciente del periodo
        planilla = PlanillaCalculo.objects.filter(periodo=periodo).order_by('-generada_en').first()
        if planilla is None:
            return Response(
                {'error': f'No existe planilla calculada para el periodo "{periodo}".'},
                status=status.HTTP_404_NOT_FOUND
            )

        from django.db.models import Sum, Count

        resumen_eps = (
            DetalleCalculo.objects
            .filter(planilla=planilla, estado_cruce='OK')
            .values('eps')
            .annotate(
                num_empleados=Count('id'),
                total_empresa=Sum('valor_empresa'),
                total_empleado=Sum('valor_empleado'),
                total_general=Sum('total_familia'),
                apoyo_no_gravable=Sum('apoyo_no_gravable'),
                apoyo_gravable=Sum('apoyo_gravable'),
            )
            .order_by('eps')
        )

        resultado = []
        for row in resumen_eps:
            resultado.append({
                'eps': row['eps'],
                'num_empleados': row['num_empleados'],
                'total_empresa': float(row['total_empresa'] or 0),
                'total_empleado': float(row['total_empleado'] or 0),
                'total_general': float(row['total_general'] or 0),
                'apoyo_no_gravable': float(row['apoyo_no_gravable'] or 0),
                'apoyo_gravable': float(row['apoyo_gravable'] or 0),
            })

        return Response({
            'periodo': periodo,
            'planilla_id': planilla.id,
            'generada_en': planilla.generada_en,
            'por_eps': resultado,
            'totales': {
                'total_empresa': float(planilla.total_empresa),
                'total_empleado': float(planilla.total_empleado),
                'total_gravable': float(planilla.total_gravable),
                'total_no_gravable': float(planilla.total_no_gravable),
                'total_empleados': planilla.total_empleados,
            },
        })


class ConciliacionView(APIView):
    """
    GET /api/beneficios-salud/conciliacion/?periodo_nuevo=X&periodo_anterior=Y
    Compara dos planillas (por periodo) y retorna novedades: nuevos, retirados, cambios de valor.
    Usa las planillas más recientes de cada periodo.
    """

    def get(self, request, *args, **kwargs):
        periodo_nuevo = request.query_params.get('periodo_nuevo')
        periodo_anterior = request.query_params.get('periodo_anterior')

        if not periodo_nuevo or not periodo_anterior:
            return Response(
                {'error': 'Se requieren los parámetros "periodo_nuevo" y "periodo_anterior".'},
                status=status.HTTP_400_BAD_REQUEST
            )

        planilla_nueva = PlanillaCalculo.objects.filter(periodo=periodo_nuevo).order_by('-generada_en').first()
        planilla_anterior = PlanillaCalculo.objects.filter(periodo=periodo_anterior).order_by('-generada_en').first()

        if planilla_nueva is None:
            return Response(
                {'error': f'No existe planilla para el periodo nuevo "{periodo_nuevo}".'},
                status=status.HTTP_404_NOT_FOUND
            )
        if planilla_anterior is None:
            return Response(
                {'error': f'No existe planilla para el periodo anterior "{periodo_anterior}".'},
                status=status.HTTP_404_NOT_FOUND
            )

        detalles_nuevos = {
            d['cedula']: d
            for d in planilla_nueva.detalles.filter(estado_cruce='OK').values(
                'cedula', 'nombre_en_kactus', 'eps', 'total_familia', 'valor_empresa'
            )
        }
        detalles_anteriores = {
            d['cedula']: d
            for d in planilla_anterior.detalles.filter(estado_cruce='OK').values(
                'cedula', 'nombre_en_kactus', 'eps', 'total_familia', 'valor_empresa'
            )
        }

        cedulas_nuevas = set(detalles_nuevos.keys())
        cedulas_anteriores = set(detalles_anteriores.keys())

        nuevos = [
            {
                'cedula': c,
                'nombre': detalles_nuevos[c]['nombre_en_kactus'],
                'eps': detalles_nuevos[c]['eps'],
                'total_familia': float(detalles_nuevos[c]['total_familia'] or 0),
            }
            for c in sorted(cedulas_nuevas - cedulas_anteriores)
        ]

        retirados = [
            {
                'cedula': c,
                'nombre': detalles_anteriores[c]['nombre_en_kactus'],
                'eps': detalles_anteriores[c]['eps'],
                'total_familia': float(detalles_anteriores[c]['total_familia'] or 0),
            }
            for c in sorted(cedulas_anteriores - cedulas_nuevas)
        ]

        cambios_valor = []
        sin_cambios = 0
        for c in cedulas_nuevas & cedulas_anteriores:
            vn = float(detalles_nuevos[c]['valor_empresa'] or 0)
            va = float(detalles_anteriores[c]['valor_empresa'] or 0)
            if abs(vn - va) > 1.0:
                cambios_valor.append({
                    'cedula': c,
                    'nombre': detalles_nuevos[c]['nombre_en_kactus'],
                    'eps': detalles_nuevos[c]['eps'],
                    'valor_anterior': va,
                    'valor_nuevo': vn,
                    'diferencia': round(vn - va, 2),
                })
            else:
                sin_cambios += 1

        return Response({
            'periodo_nuevo': periodo_nuevo,
            'periodo_anterior': periodo_anterior,
            'resumen': {
                'nuevos': len(nuevos),
                'retirados': len(retirados),
                'cambios_valor': len(cambios_valor),
                'sin_cambios': sin_cambios,
            },
            'nuevos': nuevos,
            'retirados': retirados,
            'cambios_valor': cambios_valor,
        })


class InformeEFRView(APIView):
    """
    GET /api/beneficios-salud/informe-efr/?periodo=MMYYYY
    Informe mensual EFR combinando:
      - Datos de planilla (si existe para el periodo)
      - Pensionados activos
      - Auxilio externo activo
    """

    def get(self, request, *args, **kwargs):
        from django.db.models import Sum, Count

        periodo = request.query_params.get('periodo')
        if not periodo:
            return Response({'error': 'El parámetro "periodo" es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        # Planilla
        planilla = PlanillaCalculo.objects.filter(periodo=periodo).order_by('-generada_en').first()
        planilla_resumen = None
        por_eps = []

        if planilla:
            planilla_resumen = {
                'planilla_id': planilla.id,
                'periodo': planilla.periodo,
                'total_empleados': planilla.total_empleados,
                'total_empresa': float(planilla.total_empresa),
                'total_empleado': float(planilla.total_empleado),
                'total_gravable': float(planilla.total_gravable),
                'total_no_gravable': float(planilla.total_no_gravable),
                'generada_en': planilla.generada_en,
                'generada_por': planilla.generada_por,
            }

            por_eps_qs = (
                DetalleCalculo.objects
                .filter(planilla=planilla, estado_cruce='OK')
                .values('eps')
                .annotate(
                    num_empleados=Count('id'),
                    total_empresa=Sum('valor_empresa'),
                    total_empleado=Sum('valor_empleado'),
                    apoyo_no_gravable=Sum('apoyo_no_gravable'),
                    apoyo_gravable=Sum('apoyo_gravable'),
                )
                .order_by('eps')
            )
            por_eps = [
                {
                    'eps': row['eps'],
                    'num_empleados': row['num_empleados'],
                    'total_empresa': float(row['total_empresa'] or 0),
                    'total_empleado': float(row['total_empleado'] or 0),
                    'apoyo_no_gravable': float(row['apoyo_no_gravable'] or 0),
                    'apoyo_gravable': float(row['apoyo_gravable'] or 0),
                }
                for row in por_eps_qs
            ]

        # Pensionados activos
        pensionados_qs = PensionadoPrepagada.objects.filter(activo=True)
        pensionados = PensionadoPrepagadaSerializer(pensionados_qs, many=True).data
        total_pensionados = float(
            pensionados_qs.aggregate(total=Sum('valor_mensual'))['total'] or 0
        )

        # Auxilio externo activo
        auxilio_qs = AuxilioExterno.objects.filter(activo=True)
        auxilio_externo = AuxilioExternoSerializer(auxilio_qs, many=True).data
        total_auxilio = float(
            auxilio_qs.aggregate(total=Sum('valor_mensual'))['total'] or 0
        )

        # Totales consolidados
        total_empresa = float(planilla.total_empresa) if planilla else 0
        total_empleado = float(planilla.total_empleado) if planilla else 0

        return Response({
            'periodo': periodo,
            'planilla_resumen': planilla_resumen,
            'por_eps': por_eps,
            'pensionados': {
                'lista': pensionados,
                'total': len(pensionados),
                'total_valor': total_pensionados,
            },
            'auxilio_externo': {
                'lista': auxilio_externo,
                'total': len(auxilio_externo),
                'total_valor': total_auxilio,
            },
            'consolidado': {
                'total_empresa': round(total_empresa + total_pensionados, 2),
                'total_empleado': round(total_empleado, 2),
                'total_general': round(total_empresa + total_empleado + total_pensionados + total_auxilio, 2),
            },
        })
